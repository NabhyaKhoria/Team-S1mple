from urllib import request
from django.shortcuts import redirect, render
from .models import *
import openpyxl
import requests


# Create your views here.
def index(request):
    notices = Notice.objects.all()
    supervisers = Superviser.objects.all()
    print(supervisers)
    context = {
        'notices':notices,

    }
    return render(request,'index.html',context)

def notice_detail(request,pk):
    notice = Notice.objects.filter(id=pk)
    supervisers = Superviser.objects.all()
    print(notice)
    context = {
        'notice':notice,
        'supervisers':supervisers,
    }
    return render(request,'portfolio-details.html',context)

def complaint_form(request):
    if(request.method == 'POST'):
        name = request.POST.get('name',False)
        email = request.POST.get('email',False)
        subject = request.POST.get('subject',False)
        message = request.POST.get('message',False)
        url = 'https://docs.google.com/forms/d/e/1FAIpQLSf50jP5d8fJhwvVLJptcMf1mzxAGWMPEUg8Go5KqVquKcZb9Q/formResponse'
        obj = {"entry.777368475":name,"entry.1558772464":email,"entry.220954102":subject,"entry.229301749":message,"fvv":1,"draftResponse":'[]',"pageHistory":0,"fbzx":-2525148146002960632}
        x = requests.post(url, data=obj)
        wb_obj = openpyxl.load_workbook("Complaint_Form.xlsx")
        print(wb_obj)
        sheet_obj = wb_obj.active
        j=1
        current_row=sheet_obj.max_row+1
        cell_obj = sheet_obj.cell(row = current_row, column = j)
        cell_obj.value=current_row
        j=2
        cell_obj = sheet_obj.cell(row = current_row, column = j)
        cell_obj.value=name
        j=3
        cell_obj = sheet_obj.cell(row = current_row, column = j)
        cell_obj.value=email
        j=4
        cell_obj = sheet_obj.cell(row = current_row, column = j)
        cell_obj.value=subject
        j=5
        cell_obj = sheet_obj.cell(row = current_row, column = j)
        cell_obj.value=message
        # wb_obj.save('Complaint_Form.xlsx')
        try:
            wb_obj.save('Complaint_Form.xlsx')
        except:
            print("Failure")
        print("MESSAGE =======",x)
        return redirect('mmm:home')


